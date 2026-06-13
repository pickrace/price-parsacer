"""
Запис результатів в Excel (історія накопичується: кожен запуск = новий рядок).

Схема свідомо повторює логіку майстер-файлу Gorgany:
  Дата | Запит | Бренд | {Конкурент} Ціна | {Конкурент} Посилання | ...
щоб результат лягав у вже існуючий процес. Колонки конкурентів
будуються автоматично з ACTIVE_SCRAPERS, тож при додаванні конкурента
схема розширюється сама.
"""

from __future__ import annotations

import datetime as dt
from pathlib import Path

from openpyxl import Workbook, load_workbook

from scrapers import ACTIVE_SCRAPERS, ScrapeResult

DATA_FILE = Path(__file__).parent / "data" / "price_history.xlsx"
SHEET = "History"


def _header() -> list[str]:
    cols = ["Дата", "Запит", "Бренд"]
    for cls in ACTIVE_SCRAPERS:
        cols += [f"{cls.name} Ціна", f"{cls.name} Посилання"]
    return cols


def _ensure_workbook() -> None:
    if DATA_FILE.exists():
        return
    DATA_FILE.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    ws.append(_header())
    wb.save(DATA_FILE)


def append_results(query: str, brand: str, results: list[ScrapeResult]) -> Path:
    """Додає один рядок з результатами поточного запуску. Повертає шлях до файлу."""
    _ensure_workbook()
    wb = load_workbook(DATA_FILE)
    ws = wb[SHEET]

    by_name = {r.competitor: r for r in results}
    row = [dt.date.today().isoformat(), query, brand]
    for cls in ACTIVE_SCRAPERS:
        r = by_name.get(cls.name)
        # порожньо замість "not found", щоб історія цін була чистою для аналізу
        row.append(r.price if (r and r.status == "ok") else None)
        row.append(r.url if (r and r.url) else None)
    ws.append(row)

    wb.save(DATA_FILE)
    return DATA_FILE
